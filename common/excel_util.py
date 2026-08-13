# -*- coding: utf-8 -*-
"""
Excel读写工具类
"""
import os
import json
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from config.settings import COLUMNS, EXCEL_PATH, SHEET_NAME


class ExcelUtil:
    """Excel工具类"""

    def __init__(self, excel_path=None, sheet_name=None):
        self.excel_path = excel_path or EXCEL_PATH
        self.sheet_name = sheet_name or SHEET_NAME
        self._ensure_dir()
        self._ensure_file()
        self.wb = None
        self.ws = None

    def _ensure_dir(self):
        """确保目录存在"""
        dir_path = os.path.dirname(self.excel_path)
        if not os.path.exists(dir_path):
            os.makedirs(dir_path)

    def _ensure_file(self):
        """确保Excel文件存在，不存在则创建模板"""
        if not os.path.exists(self.excel_path):
            self._create_template()

    def _create_template(self):
        """创建Excel模板文件，按用户提供的格式"""
        wb = Workbook()
        ws = wb.active
        ws.title = self.sheet_name

        # 表头 - 严格按照用户提供的图片格式
        headers = [
            "接口名称",      # A
            "接口描述",      # B
            "url",           # C
            "是否运行",      # D
            "请求类型",      # E
            "是否携带header",# F
            "case依赖",      # G
            "依赖返回数据的层级",  # H
            "依赖返回数据的key",   # I
            "依赖的返回数据",      # J
            "数据依赖字段",        # K
            "请求数据",      # L
            "预期结果",      # M
            "实际结果"       # N
        ]

        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # 示例数据 - 按用户图片中的内容
        sample_data = [
            [
                "getNewLotteryNo",
                "根据彩种ID获取当前可购彩期号",
                "http://172.16.10.221/common/getNewLotteryNo",
                "yes",
                "get",
                "no",
                "",
                "",
                "",
                "",
                "",
                json.dumps({"id": "20"}, ensure_ascii=False),
                json.dumps({"code": 20}, ensure_ascii=False),
                ""
            ],
            [
                "trade/payment",
                "购彩接口",
                "http://172.16.10.221/trade/payment",
                "yes",
                "post",
                "yes",
                "getNewLotteryResult",
                "6",
                "nextNo",
                "",
                "lotteryNo",
                json.dumps({"trade/payn": "code"}, ensure_ascii=False),
                "20",
                ""
            ],
            [
                "getNewLotteryResult",
                "获取彩种开奖结果信息",
                "http://172.16.10.221/common/getNewLotteryResult",
                "no",
                "post",
                "no",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                ""
            ]
        ]

        for row_idx, row_data in enumerate(sample_data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # 调整列宽
        col_widths = [18, 28, 45, 10, 10, 15, 20, 18, 18, 16, 16, 30, 20, 20]
        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = width

        # 冻结首行
        ws.freeze_panes = "A2"

        wb.save(self.excel_path)

    def _load_workbook(self):
        """加载工作簿"""
        self.wb = load_workbook(self.excel_path)
        if self.sheet_name in self.wb.sheetnames:
            self.ws = self.wb[self.sheet_name]
        else:
            self.ws = self.wb.active

    def read_all_cases(self):
        """
        读取所有测试用例
        :return: 测试用例列表，每个用例为字典
        """
        self._load_workbook()
        cases = []
        # 从第2行开始读取（跳过表头）
        for row_idx in range(2, self.ws.max_row + 1):
            case = self._read_row(row_idx)
            if case and case.get("interface_name"):
                case["row_num"] = row_idx
                cases.append(case)
        return cases

    def _read_row(self, row_num):
        """读取一行数据"""
        case = {}
        for key, col_idx in COLUMNS.items():
            cell = self.ws.cell(row=row_num, column=col_idx + 1)
            value = cell.value
            case[key] = value if value is not None else ""
        return case

    def write_actual_result(self, row_num, result, is_pass=None):
        """
        写入实际结果
        :param row_num: 行号（从1开始）
        :param result: 实际结果内容
        :param is_pass: 是否通过，True/False/None
        """
        self._load_workbook()
        cell = self.ws.cell(row=row_num, column=COLUMNS["actual_result"] + 1, value=str(result))

        # 设置颜色
        if is_pass is True:
            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            cell.font = Font(color="006100")
        elif is_pass is False:
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            cell.font = Font(color="9C0006")

        self.wb.save(self.excel_path)

    def write_cell(self, row_num, column_key, value):
        """
        写入指定单元格
        :param row_num: 行号
        :param column_key: 列名key，对应COLUMNS配置
        :param value: 写入值
        """
        self._load_workbook()
        if column_key in COLUMNS:
            self.ws.cell(row=row_num, column=COLUMNS[column_key] + 1, value=str(value))
            self.wb.save(self.excel_path)

    def get_case_by_name(self, case_name):
        """
        根据接口名称获取用例行号和数据
        :param case_name: 接口名称（case依赖字段对应的值）
        :return: (row_num, case_dict) 或 (None, None)
        """
        self._load_workbook()
        for row_idx in range(2, self.ws.max_row + 1):
            cell_value = self.ws.cell(row=row_idx, column=COLUMNS["interface_name"] + 1).value
            if cell_value == case_name:
                case = self._read_row(row_idx)
                case["row_num"] = row_idx
                return row_idx, case
        return None, None
